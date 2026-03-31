"""
XLSX 파일을 파싱하여 Markdown 표로 변환하고 data/markdown/에 저장하는 스크립트.

extractor.py의 파이프 구분 출력과 별도로, 마크다운 표 형식으로
사람이 읽기 좋은 결과물을 생성합니다.

실행 방법:
    python src/extract_xlsx.py
"""

import sys
import time
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

from extractor import extract_from_xlsx

# 기본 경로
BASE_DIR = Path(__file__).parent.parent
DOCS_DIR = BASE_DIR / "data" / "docs"
MARKDOWN_DIR = BASE_DIR / "data" / "markdown"


def _xlsx_to_markdown_tables(file_path: Path) -> list[dict]:
    """XLSX 파일을 시트별 마크다운 표로 변환합니다.

    빈 셀을 보존하고 열 너비를 맞추어 마크다운 표 형식으로 변환합니다.

    Args:
        file_path: XLSX 파일 경로

    Returns:
        [{"sheet_name": 시트명, "table_md": 마크다운 표 문자열}, ...]
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 모든 행을 문자열 리스트로 수집 (빈 셀은 빈 문자열 유지)
        rows = []
        for row in ws.iter_rows(max_col=ws.max_column):
            cells = []
            for cell in row:
                val = cell.value
                if val is None:
                    cells.append("")
                else:
                    cells.append(str(val).strip())
            rows.append(cells)

        # 완전히 빈 행 제거
        rows = [r for r in rows if any(c for c in r)]
        if not rows:
            continue

        # 열 너비 계산 (최소 3자)
        max_col = max(len(r) for r in rows)
        # 열 개수 통일
        for r in rows:
            while len(r) < max_col:
                r.append("")

        # 마크다운 표 생성
        md_lines = []

        # 첫 번째 행 = 헤더
        header = rows[0]
        md_lines.append("| " + " | ".join(header) + " |")
        md_lines.append("| " + " | ".join(["---"] * max_col) + " |")

        # 나머지 행 = 데이터
        for row in rows[1:]:
            md_lines.append("| " + " | ".join(row) + " |")

        sheets.append({
            "sheet_name": sheet_name,
            "table_md": "\n".join(md_lines),
        })

    return sheets


def save_as_markdown(file_path: Path, result: dict, output_dir: Path) -> Path:
    """추출 결과를 마크다운 표 형식으로 저장합니다.

    Args:
        file_path: 원본 XLSX 파일 경로 (마크다운 표 생성용)
        result: extractor.extract_from_xlsx() 반환값 (메타정보용)
        output_dir: Markdown 파일 저장 디렉토리

    Returns:
        저장된 Markdown 파일 경로
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = Path(result["file_name"]).stem
    md_path = output_dir / f"{stem}.md"

    # 마크다운 표 생성
    sheets = _xlsx_to_markdown_tables(file_path)

    lines = []
    lines.append(f"# {result['file_name']}\n")
    lines.append(f"- 파일 형식: {result['file_type']}")
    lines.append(f"- 총 시트: {len(result['pages'])}개")
    lines.append(f"- 추출 글자 수: {len(result['full_text'])}자\n")
    lines.append("---\n")

    for sheet in sheets:
        lines.append(f"## 시트: {sheet['sheet_name']}\n")
        lines.append(sheet["table_md"])
        lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")
    return md_path


def main() -> None:
    """data/docs/에서 모든 XLSX를 찾아 파싱하고 Markdown으로 저장합니다."""
    print("\n" + "=" * 60)
    print("  📝 XLSX 파싱 → Markdown 표 변환")
    print("=" * 60)

    # XLSX 파일 수집
    xlsx_files = sorted(DOCS_DIR.rglob("*.xlsx"))
    if not xlsx_files:
        print(f"  ❌ XLSX 파일이 없습니다: {DOCS_DIR}")
        sys.exit(1)

    print(f"  📁 문서: data/docs/")
    print(f"  📄 XLSX 파일: {len(xlsx_files)}개\n")

    start_time = time.time()
    results = []

    for file_path in xlsx_files:
        result = extract_from_xlsx(file_path)
        md_path = save_as_markdown(file_path, result, MARKDOWN_DIR)
        text_len = len(result["full_text"])
        sheet_count = len(result["pages"])
        print(f"  ✅ {file_path.name} → {md_path.name} ({sheet_count}시트, {text_len}자)")
        results.append(result)

    elapsed = time.time() - start_time
    print(f"\n  ✅ XLSX 파싱 완료: {len(results)}개 ({elapsed:.1f}초)")
    print(f"  💾 저장 위치: data/markdown/")


if __name__ == "__main__":
    main()
